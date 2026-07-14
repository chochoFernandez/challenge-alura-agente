"""Etapa 2 del pipeline: extracción, limpieza y chunking de los documentos.

De acá salen los Chunks que después se embeben e indexan. Si esta etapa produce basura,
ningún prompt ni modelo la salva después: la citación de fuentes depende enteramente de
los metadatos que se adjuntan acá.

Cada formato se extrae distinto:

    PDF       texto por página (pypdf), descartando encabezados y pies repetidos
    DOCX      párrafos y títulos, agrupando por sección
    CSV/XLSX  fila por fila, repitiendo los encabezados de columna en cada fila
    JSON      aplanado a frases legibles ("plan Cuenta Plus > precio_mensual: 4900")
    Markdown  se corta por encabezado, preservando la jerarquía de secciones
"""

import csv
import hashlib
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import get_settings

logger = logging.getLogger(__name__)

# Qué categoría de negocio le corresponde a cada documento. En una empresa real esto
# saldría de un catálogo con responsables por área; acá lo declaramos a mano porque son
# seis archivos. Es el metadato que permitiría filtrar búsquedas por área.
CATEGORIAS = {
    "politica_rrhh.pdf": "Recursos Humanos",
    "politica_privacidad.docx": "Legal y Compliance",
    "tarifas_comisiones.csv": "Financiero",
    "planes_productos.json": "Comercial",
    "faq_transacciones.md": "Operacional",
    "ventas_2025.xlsx": "Datos y Sistemas",
}


@dataclass
class Chunk:
    """Un fragmento de documento listo para indexar, con su procedencia."""

    texto: str
    source_file: str
    category: str
    location: str  # "página 2", "sección 3. Licencia...", "hoja Resumen Mensual, fila 48"
    chunk_id: str = field(default="")

    def __post_init__(self) -> None:
        if not self.chunk_id:
            # Determinístico: mismo contenido + misma procedencia => mismo id.
            # Es la base de la idempotencia del índice (reconstruirlo no duplica nada).
            crudo = f"{self.source_file}|{self.location}|{self.texto}"
            self.chunk_id = hashlib.sha256(crudo.encode("utf-8")).hexdigest()[:16]

    @property
    def cita(self) -> str:
        """Cómo se muestra esta fuente al usuario."""
        return f"{self.source_file} ({self.location})"


# --------------------------------------------------------------------------------------
# Limpieza
# --------------------------------------------------------------------------------------


def limpiar_texto(texto: str) -> str:
    """Normaliza espacios y saltos de línea sin alterar el contenido."""
    texto = texto.replace("\xa0", " ")  # espacio duro
    texto = re.sub(r"[ \t]+", " ", texto)  # espacios repetidos
    texto = re.sub(r"\n{3,}", "\n\n", texto)  # más de un renglón en blanco
    texto = re.sub(r"(?<=\w)-\n(?=\w)", "", texto)  # palabra cortada con guion al fin de línea
    return texto.strip()


def detectar_lineas_repetidas(paginas: list[str], umbral: float = 0.6) -> set[str]:
    """Detecta encabezados y pies de página: líneas que se repiten en casi todas las páginas.

    En un PDF corporativo, "Documento interno confidencial" y "Página 3" aparecen en cada
    página. Son ruido: no aportan significado y ensucian los embeddings. En vez de
    hardcodear qué líneas descartar, se detectan por frecuencia — así funciona con
    cualquier PDF, no solo con el nuestro.
    """
    if len(paginas) < 2:
        return set()

    conteo: dict[str, int] = {}
    for pagina in paginas:
        # Solo miramos el principio y el final de cada página: ahí viven encabezados y pies.
        # La ventana es de 3 líneas y no de 2 porque un pie suele traer varias (el título del
        # documento y el número de página, por ejemplo), y con una ventana corta la última
        # se escapa y termina contaminando un chunk.
        lineas = [ln.strip() for ln in pagina.splitlines() if ln.strip()]
        for linea in lineas[:3] + lineas[-3:]:
            # "Página 3" y "Página 4" son la misma línea a estos efectos.
            normalizada = re.sub(r"\d+", "#", linea)
            conteo[normalizada] = conteo.get(normalizada, 0) + 1

    minimo = max(2, int(len(paginas) * umbral))
    return {linea for linea, veces in conteo.items() if veces >= minimo}


def _quitar_lineas_repetidas(pagina: str, repetidas: set[str]) -> str:
    conservadas = [
        linea
        for linea in pagina.splitlines()
        if re.sub(r"\d+", "#", linea.strip()) not in repetidas
    ]
    return "\n".join(conservadas)


# --------------------------------------------------------------------------------------
# Chunking
# --------------------------------------------------------------------------------------


def _cola_con_overlap(texto: str, overlap: int) -> str:
    """Devuelve los últimos ~overlap caracteres, sin cortar una palabra por la mitad.

    Cortar a ciegas produce colas como "a de esa franja cada equipo...", que arrancan con
    un fragmento de palabra sin sentido y ensucian el embedding del chunk siguiente.
    """
    if overlap <= 0 or not texto:
        return ""

    cola = texto[-overlap:]
    corte = cola.find(" ")
    if corte != -1:
        cola = cola[corte + 1 :]
    return cola.strip()


def partir_en_chunks(texto: str, tam: int, overlap: int) -> list[str]:
    """Parte un texto en fragmentos de ~tam caracteres, cortando en límites de oración.

    Cortar a ciegas cada N caracteres parte ideas por la mitad ("...tiene derecho a 22" |
    "días hábiles..."), lo que degrada el embedding del fragmento. Acá se acumulan párrafos
    y oraciones enteras hasta llegar al tamaño objetivo.
    """
    texto = texto.strip()
    if not texto:
        return []
    if len(texto) <= tam:
        return [texto]

    # Unidades atómicas: párrafos, y si un párrafo ya excede el tamaño, sus oraciones.
    unidades: list[str] = []
    for parrafo in re.split(r"\n\s*\n", texto):
        parrafo = parrafo.strip()
        if not parrafo:
            continue
        if len(parrafo) <= tam:
            unidades.append(parrafo)
        else:
            unidades.extend(o.strip() for o in re.split(r"(?<=[.!?])\s+", parrafo) if o.strip())

    chunks: list[str] = []
    actual = ""
    for unidad in unidades:
        candidato = f"{actual}\n\n{unidad}" if actual else unidad
        if len(candidato) <= tam:
            actual = candidato
            continue

        if actual:
            chunks.append(actual)
            # Overlap: arrastramos la cola del chunk anterior para no perder el hilo de una
            # idea que quedó a caballo entre dos fragmentos.
            cola = _cola_con_overlap(actual, overlap)
            actual = f"{cola}\n\n{unidad}" if cola else unidad
        else:
            actual = unidad

        # Una sola unidad más larga que el tamaño objetivo: la cortamos duro, no hay opción.
        while len(actual) > tam:
            chunks.append(actual[:tam])
            actual = actual[tam - overlap :]

    if actual.strip():
        chunks.append(actual.strip())

    return chunks


# --------------------------------------------------------------------------------------
# Extractores por formato
# --------------------------------------------------------------------------------------


def _cargar_pdf(ruta: Path, categoria: str, tam: int, overlap: int) -> list[Chunk]:
    lector = PdfReader(ruta)
    paginas = [pagina.extract_text() or "" for pagina in lector.pages]
    repetidas = detectar_lineas_repetidas(paginas)
    if repetidas:
        logger.info("%s: descartando %d línea(s) de encabezado/pie", ruta.name, len(repetidas))

    chunks: list[Chunk] = []
    for numero, pagina in enumerate(paginas, start=1):
        texto = limpiar_texto(_quitar_lineas_repetidas(pagina, repetidas))
        for fragmento in partir_en_chunks(texto, tam, overlap):
            chunks.append(
                Chunk(
                    texto=fragmento,
                    source_file=ruta.name,
                    category=categoria,
                    location=f"página {numero}",
                )
            )
    return chunks


def _cargar_docx(ruta: Path, categoria: str, tam: int, overlap: int) -> list[Chunk]:
    doc = Document(ruta)

    # Agrupamos por sección: el texto de cada título viaja junto a sus párrafos, así el
    # fragmento conserva de qué habla ("4. Derechos del titular" + su contenido).
    secciones: list[tuple[str, list[str]]] = []
    titulo_actual = "Encabezado"
    cuerpo: list[str] = []

    for parrafo in doc.paragraphs:
        texto = parrafo.text.strip()
        if not texto:
            continue
        if parrafo.style.name.startswith("Heading") or parrafo.style.name == "Title":
            if cuerpo:
                secciones.append((titulo_actual, cuerpo))
            titulo_actual = texto
            cuerpo = []
        else:
            cuerpo.append(texto)
    if cuerpo:
        secciones.append((titulo_actual, cuerpo))

    chunks: list[Chunk] = []
    for titulo, parrafos in secciones:
        texto = limpiar_texto(f"{titulo}\n\n" + "\n\n".join(parrafos))
        for fragmento in partir_en_chunks(texto, tam, overlap):
            chunks.append(
                Chunk(
                    texto=fragmento,
                    source_file=ruta.name,
                    category=categoria,
                    location=f"sección «{titulo}»",
                )
            )
    return chunks


def _fila_a_texto(encabezados: list[str], valores: list) -> str:
    """Convierte una fila tabular en una frase legible, repitiendo los encabezados.

    Una fila suelta ("Extracción en cajero | Efectivo | 850 | ARS") no significa nada sin
    sus columnas: el embedding sería inútil. Repetir el encabezado en cada fila es lo que
    permite que "¿cuánto cuesta sacar plata del cajero?" recupere esta fila.
    """
    partes = [
        f"{col}: {val}"
        for col, val in zip(encabezados, valores)
        if val is not None and str(val).strip() != ""
    ]
    return " | ".join(partes)


def _cargar_csv(ruta: Path, categoria: str, tam: int, overlap: int) -> list[Chunk]:
    with ruta.open(encoding="utf-8", newline="") as f:
        filas = list(csv.reader(f))

    if not filas:
        logger.warning("%s está vacío, no genera chunks", ruta.name)
        return []

    encabezados, datos = filas[0], filas[1:]
    chunks: list[Chunk] = []
    for numero, fila in enumerate(datos, start=2):  # fila 1 = encabezado
        texto = _fila_a_texto(encabezados, fila)
        if not texto:
            continue
        chunks.append(
            Chunk(
                texto=texto,
                source_file=ruta.name,
                category=categoria,
                location=f"fila {numero}",
            )
        )
    return chunks


def _cargar_xlsx(ruta: Path, categoria: str, tam: int, overlap: int) -> list[Chunk]:
    wb = load_workbook(ruta, read_only=True, data_only=True)
    chunks: list[Chunk] = []

    for hoja in wb.worksheets:
        filas = list(hoja.iter_rows(values_only=True))
        if not filas:
            continue

        encabezados = [str(c) if c is not None else "" for c in filas[0]]
        for numero, fila in enumerate(filas[1:], start=2):
            texto = _fila_a_texto(encabezados, list(fila))
            if not texto:
                continue
            chunks.append(
                Chunk(
                    texto=texto,
                    source_file=ruta.name,
                    category=categoria,
                    location=f"hoja «{hoja.title}», fila {numero}",
                )
            )

    wb.close()
    return chunks


def _aplanar_json(dato, prefijo: str = "") -> list[str]:
    """Convierte un JSON anidado en frases planas del tipo 'a > b > c: valor'."""
    frases: list[str] = []

    if isinstance(dato, dict):
        for clave, valor in dato.items():
            nuevo = f"{prefijo} > {clave}" if prefijo else str(clave)
            frases.extend(_aplanar_json(valor, nuevo))
    elif isinstance(dato, list):
        for i, item in enumerate(dato):
            # Si el item tiene nombre propio, lo usamos como etiqueta en vez del índice:
            # "planes > Cuenta Plus > precio" se lee mucho mejor que "planes > 1 > precio".
            etiqueta = item.get("nombre") if isinstance(item, dict) and item.get("nombre") else str(i)
            frases.extend(_aplanar_json(item, f"{prefijo} > {etiqueta}" if prefijo else etiqueta))
    else:
        valor = "sin especificar" if dato is None else str(dato)
        frases.append(f"{prefijo}: {valor}")

    return frases


def _cargar_json(ruta: Path, categoria: str, tam: int, overlap: int) -> list[Chunk]:
    dato = json.loads(ruta.read_text(encoding="utf-8"))

    # Un chunk por elemento de primer nivel de la lista principal (cada plan por separado):
    # así preguntar por el plan Business no arrastra los otros tres planes al contexto.
    chunks: list[Chunk] = []

    if isinstance(dato, dict) and isinstance(dato.get("planes"), list):
        contexto = {k: v for k, v in dato.items() if k != "planes"}
        cabecera = " | ".join(_aplanar_json(contexto))

        for plan in dato["planes"]:
            nombre = plan.get("nombre", plan.get("id", "plan"))
            cuerpo = "\n".join(_aplanar_json(plan, nombre))
            texto = limpiar_texto(f"{cabecera}\n\n{cuerpo}")
            for fragmento in partir_en_chunks(texto, tam, overlap):
                chunks.append(
                    Chunk(
                        texto=fragmento,
                        source_file=ruta.name,
                        category=categoria,
                        location=f"plan «{nombre}»",
                    )
                )
        return chunks

    # JSON de forma desconocida: lo aplanamos entero y lo partimos por tamaño.
    texto = limpiar_texto("\n".join(_aplanar_json(dato)))
    return [
        Chunk(texto=fragmento, source_file=ruta.name, category=categoria, location="documento")
        for fragmento in partir_en_chunks(texto, tam, overlap)
    ]


def _cargar_markdown(ruta: Path, categoria: str, tam: int, overlap: int) -> list[Chunk]:
    texto = ruta.read_text(encoding="utf-8")

    # Cortamos por encabezado: cada sección de la FAQ es una unidad de sentido completa
    # (una pregunta y su respuesta), que es justamente lo que queremos recuperar.
    partes = re.split(r"^(#{1,6}\s+.*)$", texto, flags=re.MULTILINE)

    secciones: list[tuple[str, str]] = []
    intro = partes[0].strip()
    if intro:
        secciones.append(("Introducción", intro))
    for i in range(1, len(partes), 2):
        titulo = re.sub(r"^#+\s+", "", partes[i]).strip()
        cuerpo = partes[i + 1] if i + 1 < len(partes) else ""
        secciones.append((titulo, f"{titulo}\n\n{cuerpo.strip()}"))

    chunks: list[Chunk] = []
    for titulo, cuerpo in secciones:
        limpio = limpiar_texto(re.sub(r"\*\*(.+?)\*\*", r"\1", cuerpo))  # quitamos el negrita
        for fragmento in partir_en_chunks(limpio, tam, overlap):
            chunks.append(
                Chunk(
                    texto=fragmento,
                    source_file=ruta.name,
                    category=categoria,
                    location=f"sección «{titulo}»",
                )
            )
    return chunks


EXTRACTORES = {
    ".pdf": _cargar_pdf,
    ".docx": _cargar_docx,
    ".csv": _cargar_csv,
    ".xlsx": _cargar_xlsx,
    ".json": _cargar_json,
    ".md": _cargar_markdown,
}


# --------------------------------------------------------------------------------------
# Entrada principal
# --------------------------------------------------------------------------------------


def cargar_documento(ruta: Path) -> list[Chunk]:
    """Extrae, limpia y trocea un único documento."""
    settings = get_settings()
    extractor = EXTRACTORES.get(ruta.suffix.lower())
    if extractor is None:
        raise ValueError(
            f"Formato no soportado: {ruta.suffix} ({ruta.name}). "
            f"Soportados: {', '.join(sorted(EXTRACTORES))}"
        )

    categoria = CATEGORIAS.get(ruta.name, "Sin categoría")
    return extractor(ruta, categoria, settings.chunk_size, settings.chunk_overlap)


def cargar_documentos(directorio: Path | None = None) -> list[Chunk]:
    """Carga todos los documentos soportados de un directorio.

    Un archivo que falla no tumba la ingesta completa: se registra el error con su causa y
    se sigue con el resto. Si NINGÚN documento se pudo leer, ahí sí se levanta el error,
    porque indexar cero chunks es un fallo silencioso que aparecería recién al preguntar.
    """
    settings = get_settings()
    directorio = directorio or settings.docs_dir

    if not directorio.is_dir():
        raise FileNotFoundError(
            f"No existe el directorio de documentos: {directorio}\n"
            "Generá la base con: python scripts/generate_docs.py"
        )

    archivos = sorted(p for p in directorio.iterdir() if p.suffix.lower() in EXTRACTORES)
    if not archivos:
        raise FileNotFoundError(
            f"No hay documentos soportados en {directorio}. "
            f"Formatos válidos: {', '.join(sorted(EXTRACTORES))}"
        )

    todos: list[Chunk] = []
    fallidos: list[str] = []

    for archivo in archivos:
        try:
            chunks = cargar_documento(archivo)
        except Exception as e:
            # No silenciamos: dejamos constancia de qué archivo falló y por qué.
            logger.error("No se pudo procesar %s: %s", archivo.name, e, exc_info=True)
            fallidos.append(archivo.name)
            continue

        logger.info("%s: %d chunks", archivo.name, len(chunks))
        todos.extend(chunks)

    if not todos:
        raise RuntimeError(
            f"Ningún documento pudo procesarse en {directorio}. Fallaron: {', '.join(fallidos)}"
        )
    if fallidos:
        logger.warning("Se ignoraron %d documento(s) con error: %s", len(fallidos), ", ".join(fallidos))

    return todos
